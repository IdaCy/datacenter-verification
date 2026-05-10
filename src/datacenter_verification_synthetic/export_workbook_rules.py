"""Export selected workbook sheets into machine-readable JSON."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from .common import write_json
except ImportError:  # pragma: no cover - direct script execution
    from common import write_json


SHEET_TO_FILE = {
    "Ground Truth Ranges": "ground_truth_ranges.json",
    "Composite Rules": "composite_rules.json",
    "Observable Matrix": "observable_matrix.json",
    "Label Definitions": "label_definitions.json",
    "Windowing Guide": "windowing_guide.json",
    "Feature Engineering": "feature_engineering.json",
}

HEADER_ROW_HINTS = {
    "Ground Truth Ranges": "Observable ID",
    "Composite Rules": "Composite label",
    "Observable Matrix": "Observable ID",
    "Label Definitions": "Label code",
    "Windowing Guide": "Window length",
    "Feature Engineering": "Feature family",
}


def _clean_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    return value


def _clean_header(value: Any, fallback: str) -> str:
    if value is None or str(value).strip() == "":
        return fallback
    text = str(value).strip()
    text = text.replace(" / ", "_").replace(" ", "_").replace("-", "_")
    text = text.replace("__", "_").lower()
    return "".join(ch for ch in text if ch.isalnum() or ch == "_")


def _row_has_content(values: list[Any]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in values)


def _find_header_row(rows: list[list[Any]], sheet_name: str) -> int:
    hint = HEADER_ROW_HINTS.get(sheet_name)
    if hint:
        for idx, row in enumerate(rows):
            if any(str(value).strip() == hint for value in row if value is not None):
                return idx
    for idx, row in enumerate(rows):
        non_empty = sum(1 for value in row if value is not None and str(value).strip() != "")
        if non_empty >= 2:
            return idx
    return 0


def sheet_to_records(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"workbook is missing required sheet {sheet_name!r}")
    worksheet = workbook[sheet_name]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    header_idx = _find_header_row(rows, sheet_name)
    raw_headers = rows[header_idx] if rows else []
    last_header_index = 0
    for idx, value in enumerate(raw_headers):
        if value is not None and str(value).strip() != "":
            last_header_index = idx
    raw_headers = raw_headers[: last_header_index + 1]
    headers = [_clean_header(value, f"column_{idx + 1}") for idx, value in enumerate(raw_headers)]

    records: list[dict[str, Any]] = []
    for source_row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        row = row[: len(headers)]
        if not _row_has_content(row):
            continue
        record = {"source_row_number": source_row_number}
        for header, value in zip(headers, row):
            record[header] = _clean_cell(value)
        records.append(record)

    return {
        "sheet_name": sheet_name,
        "source_workbook": str(workbook_path),
        "header_row_number": header_idx + 1,
        "columns": headers,
        "row_count": len(records),
        "rows": records,
    }


def fallback_rules() -> dict[str, dict[str, Any]]:
    return {
        "ground_truth_ranges.json": {
            "sheet_name": "Ground Truth Ranges",
            "source_workbook": None,
            "row_count": 0,
            "rows": [],
            "note": "Workbook unavailable; generated fallback placeholder.",
        },
        "composite_rules.json": {
            "sheet_name": "Composite Rules",
            "source_workbook": None,
            "rows": [
                {
                    "composite_label": "0 - No training likely",
                    "suggested_rule_for_synthetic_ground_truth": "capacity below threshold or all primary activity low with strong coverage",
                },
                {
                    "composite_label": "1 - Training possible",
                    "suggested_rule_for_synthetic_ground_truth": "capacity exists or evidence too sparse for no-run",
                },
                {
                    "composite_label": "2 - Elevated training probability",
                    "suggested_rule_for_synthetic_ground_truth": "one primary layer elevated or physical anomaly with missing activity telemetry",
                },
                {
                    "composite_label": "3 - Training likely happening",
                    "suggested_rule_for_synthetic_ground_truth": "two independent primary layers or one primary plus corroboration",
                },
                {
                    "composite_label": "4 - Highest warning / definite",
                    "suggested_rule_for_synthetic_ground_truth": "authenticated ML evidence or coherent scheduler/GPU/fabric/power/storage evidence",
                },
            ],
        },
        "observable_matrix.json": {"sheet_name": "Observable Matrix", "source_workbook": None, "rows": []},
        "label_definitions.json": {"sheet_name": "Label Definitions", "source_workbook": None, "rows": []},
        "windowing_guide.json": {"sheet_name": "Windowing Guide", "source_workbook": None, "rows": []},
        "feature_engineering.json": {"sheet_name": "Feature Engineering", "source_workbook": None, "rows": []},
    }


def export_workbook_rules(workbook_path: Path, output_dir: Path, allow_fallback: bool = True) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    exported: dict[str, Any] = {}
    if workbook_path.exists():
        for sheet_name, file_name in SHEET_TO_FILE.items():
            payload = sheet_to_records(workbook_path, sheet_name)
            write_json(output_dir / file_name, payload)
            exported[file_name] = {"row_count": payload.get("row_count", 0), "source": str(workbook_path)}
    elif allow_fallback:
        for file_name, payload in fallback_rules().items():
            write_json(output_dir / file_name, payload)
            exported[file_name] = {"row_count": len(payload.get("rows", [])), "source": None}
    else:
        raise FileNotFoundError(f"workbook not found: {workbook_path}")
    return exported


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("xx_private/docs/ai_training_run_ground_truth_ranges.xlsx"),
    )
    parser.add_argument("--output", type=Path, default=Path("data/synthetic_v0/workbook_rules"))
    parser.add_argument("--no-fallback", action="store_true")
    args = parser.parse_args()
    exported = export_workbook_rules(args.workbook, args.output, allow_fallback=not args.no_fallback)
    for file_name, meta in exported.items():
        print(f"{file_name}: {meta['row_count']} rows")


if __name__ == "__main__":
    main()

